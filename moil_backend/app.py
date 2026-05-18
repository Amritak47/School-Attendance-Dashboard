"""
=============================================================================
Moil Primary School — Attendance Management System
Backend Server (app.py)
=============================================================================

Built for the Attendance Officer role at Moil Primary School, Darwin NT.

Technology stack:
  - Flask   : Python web framework — handles all routes and API endpoints
  - SQLite  : Lightweight database stored as a single file (attendance.db)
  - Pandas  : Reads and parses XLS/XLSX attendance export files
  - Jinja2  : Template engine for rendering HTML pages (built into Flask)

How it works:
  1. Attendance Officer uploads an XLS file exported from the school system
  2. parse_xls_file() reads the file and extracts each student's data
  3. Data is stored in the SQLite database under a unique upload ID
  4. The dashboard route loads students + their saved case notes/statuses
  5. All case updates (status changes, notes) are saved back to the database
  6. The database persists everything — notes survive page refresh and new uploads

Database tables:
  uploads           — each XLS file that has been uploaded
  students          — each student row parsed from an upload
  cases             — case status + notes per student (persistent across uploads)
  case_history      — append-only log of every status change
  departed_students — students who have left school (hidden from dashboards)
  case_plans        — full case management plan data per student (JSON)

To run:
  cd moil_backend
  python app.py
  Then open http://localhost:5000 in your browser

Author: Built for Moil Primary School Attendance Officer
Version: 1.0 — Term 1 2026
=============================================================================
"""

from flask import Flask, request, jsonify, render_template, render_template_string, send_file, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from functools import wraps
import sqlite3, os, json, re, tempfile, shutil, glob
from datetime import datetime, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter


# =============================================================================
# APP CONFIGURATION
# =============================================================================

app = Flask(__name__)

# Maximum upload file size: 32MB (large enough for any school XLS export)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-only-change-in-production')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)  # auto-logout after 8 hours (one school day)

# Database path — single file, easy to back up by copying
DB_PATH      = 'instance/attendance.db'
BACKUP_DIR   = 'backups'
BACKUP_KEEP  = 8  # keep 8 weekly backups (~2 months)

# Allowed file extensions for upload
ALLOWED = {'.xls', '.xlsx'}

# Create required folders if they don't exist yet
os.makedirs('uploads',    exist_ok=True)
os.makedirs('instance',   exist_ok=True)
os.makedirs(BACKUP_DIR,   exist_ok=True)


# =============================================================================
# FLASK-LOGIN SETUP
# =============================================================================

login_manager = LoginManager(app)
login_manager.login_view = 'login'          # redirect here when @login_required fails
login_manager.login_message = ''            # suppress default flash message (we handle it in template)


class User(UserMixin):
    """Lightweight user object loaded from the users table for Flask-Login."""
    def __init__(self, id, username, role, display_name, form_access=None):
        self.id           = id
        self.username     = username
        self.role         = role
        self.display_name = display_name
        self.form_access  = form_access  # None = all classes; 'BUSHBEES' = restricted to one class

    @property
    def is_admin(self):
        return self.role == 'admin'


@login_manager.user_loader
def load_user(user_id):
    """Load a User object from the database by ID — called by Flask-Login on each request."""
    db  = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (int(user_id),)).fetchone()
    db.close()
    if row:
        fa = row['form_access'] if 'form_access' in row.keys() else None
        return User(row['id'], row['username'], row['role'], row['display_name'], fa)
    return None


def admin_required(f):
    """Decorator: route is accessible only by admin users."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


# =============================================================================
# DATABASE SETUP
# =============================================================================

def get_db():
    """
    Open and return a connection to the SQLite database.

    row_factory = sqlite3.Row lets us access columns by name (row['name'])
    instead of by index (row[1]) — much more readable.

    Each request opens its own connection and closes it when done.
    SQLite handles concurrent reads fine; writes are serialised automatically.
    """
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")   # safer writes, better concurrency
    conn.execute("PRAGMA synchronous=NORMAL") # good balance of speed vs durability
    return conn


# Default case plan template — labels, placeholders and section titles.
# Stored in the settings table as JSON so admins can edit without touching code.
DEFAULT_CP_TEMPLATE = {
    "sections": {
        "student_info":    "Student Information",
        "student_profile": "Student Profile",
        "support":         "Support & Intervention",
        "followup":        "Follow-Up Actions",
        "signatures":      "Signatures & Sign-Off"
    },
    "fields": {
        "cp-dob":          {"label": "Date of Birth (Age)",                  "placeholder": "e.g. 12/03/2018 (8)"},
        "cp-gender":       {"label": "Gender"},
        "cp-casemanager":  {"label": "Case Manager",                         "placeholder": "Name"},
        "cp-checkin":      {"label": "Check In / Out Person",                "placeholder": "Name"},
        "cp-goal":         {"label": "Attendance Goal",                      "placeholder": "e.g. Attend 4 days per week consistently"},
        "cp-strengths":    {"label": "Strengths of Student",                 "placeholder": "Note the student's strengths, interests and positive qualities…"},
        "cp-classes":      {"label": "Student Identified Classes / Subjects","placeholder": "Subjects or activities the student enjoys or engages with…"},
        "cp-learning":     {"label": "Learning Goals",                       "placeholder": "Academic or personal learning goals for this student…"},
        "cp-barriers":     {"label": "Barriers to Attendance",               "placeholder": "What is preventing this student from attending? (transport, family, health, anxiety, bullying…)"},
        "cp-success":      {"label": "Signs of Success",                     "placeholder": "What does improvement look like for this student?"},
        "cp-rewards":      {"label": "Individual Reward System",             "placeholder": "Incentives, rewards or recognition strategies for this student…"},
        "cp-strategies":   {"label": "Strategies to Improve Attendance",     "placeholder": "List specific strategies being implemented to improve this student's attendance…"},
        "cp-sup-curriculum":{"label": "Curriculum Differentiation Plan"},
        "cp-sup-career":   {"label": "Career Counselling"},
        "cp-sup-basicneeds":{"label": "Support Basic Needs"},
        "cp-sup-mental":   {"label": "Mental Health Support"},
        "cp-sup-behaviour":{"label": "Behaviour Support Plan"},
        "cp-sup-social":   {"label": "Social Skill Development Training"},
        "cp-agency1":      {"label": "External Agency 1 (and whom)",         "placeholder": "e.g. Anglicare — Family Support Worker"},
        "cp-agency2":      {"label": "External Agency 2 (and whom)",         "placeholder": "e.g. NTCAT — Caseworker"},
        "cp-fu-notes":     {"label": "Follow-Up Notes",                      "placeholder": "Detailed notes on all follow-up contact — dates, outcomes, family responses, next steps…"}
    }
}


def get_cp_template():
    """
    Return the case plan template from the database, falling back to
    DEFAULT_CP_TEMPLATE if no custom template has been saved yet.
    """
    db  = get_db()
    row = db.execute("SELECT value FROM settings WHERE key='cp_template'").fetchone()
    db.close()
    if row:
        try:
            return json.loads(row['value'])
        except Exception:
            pass
    return DEFAULT_CP_TEMPLATE


LOGO_FILENAME = 'school_logo'   # stored in static/, extension determined at upload time

def get_school_settings():
    """Return school name and logo URL from the settings table."""
    db  = get_db()
    rows = db.execute("SELECT key, value FROM settings WHERE key IN ('school_name','school_logo_ext')").fetchall()
    db.close()
    s = {r['key']: r['value'] for r in rows}
    name     = s.get('school_name', 'My School')
    logo_ext = s.get('school_logo_ext', '')
    if logo_ext:
        logo_url = f'/static/{LOGO_FILENAME}{logo_ext}?v={os.path.getmtime("static/" + LOGO_FILENAME + logo_ext) if os.path.exists("static/" + LOGO_FILENAME + logo_ext) else 0}'
    else:
        logo_url = '/static/logo.svg'
    return {'name': name, 'logo_url': logo_url}


@app.context_processor
def inject_school():
    """Inject school name and logo URL into every Jinja2 template context."""
    return dict(school=get_school_settings())


def init_db():
    """
    Create all database tables on first run (IF NOT EXISTS = safe to call every startup).

    Schema design notes:
      - students table has one row per student PER upload
        (same student appears multiple times as new weekly data is uploaded)
      - cases table has ONE row per student across ALL uploads
        (UNIQUE on student_ref) — this is where persistent notes/status lives
      - case_plans stores the entire form as a JSON blob — easy to extend
        without database migrations if new fields are added to the form
    """
    db = get_db()
    db.executescript("""

    -- One row per uploaded XLS file
    CREATE TABLE IF NOT EXISTS uploads (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        filename      TEXT NOT NULL,           -- filename saved on disk (timestamped)
        label         TEXT,                    -- human label e.g. "Week 7 Cumulative"
        week_number   INTEGER,                 -- optional week number
        term          TEXT DEFAULT 'Term 1 2026',
        date_from     TEXT,                    -- parsed from XLS e.g. "26 JAN 2026"
        date_to       TEXT,
        upload_date   TEXT DEFAULT CURRENT_TIMESTAMP,
        student_count INTEGER,
        parsed        INTEGER DEFAULT 0        -- 1 once successfully parsed
    );

    -- One row per student per upload
    -- Sessions = number of half-days (2 sessions = 1 full school day)
    CREATE TABLE IF NOT EXISTS students (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        ref           INTEGER NOT NULL,        -- unique student ID from school system
        name          TEXT NOT NULL,           -- "Surname, Firstname" format
        year          TEXT,                    -- year level e.g. "05", "PR", "TR"
        form          TEXT,                    -- class e.g. "ACACIA", "BUSHBEES"
        upload_id     INTEGER,                 -- which upload this data came from
        attended      INTEGER,                 -- sessions attended (half-days)
        sessions      INTEGER,                 -- total sessions possible
        absences      INTEGER,                 -- sessions absent
        pct           REAL,                    -- attendance % rounded to 2dp
        days_attended REAL,                    -- attended / 2 (full days)
        days_total    REAL,                    -- sessions / 2
        days_absent   REAL,                    -- absences / 2
        FOREIGN KEY (upload_id) REFERENCES uploads(id)
    );

    -- Persistent case management — one row per student (UNIQUE on student_ref)
    -- This survives across uploads — updating student attendance data does NOT
    -- overwrite these records
    CREATE TABLE IF NOT EXISTS cases (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        student_ref   INTEGER NOT NULL,
        student_name  TEXT NOT NULL,
        form          TEXT,
        status        TEXT DEFAULT 'pending',  -- pending/contacted/meeting/welfare/referred/agency/resolved/watchlist
        notes         TEXT DEFAULT '',         -- free-text case notes
        last_updated  TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_by    TEXT DEFAULT 'Officer',
        UNIQUE(student_ref)                    -- one case record per student, ever
    );

    -- Append-only audit log — every status change gets a row added here
    -- Never updated or deleted — provides full history for accountability
    CREATE TABLE IF NOT EXISTS case_history (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        student_ref   INTEGER,
        student_name  TEXT,
        action        TEXT,                    -- e.g. "status_change"
        old_status    TEXT,
        new_status    TEXT,
        notes         TEXT,
        timestamp     TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_by    TEXT DEFAULT 'Officer'
    );

    -- Students who have left the school
    -- Hidden from all dashboard views but records remain for accountability
    CREATE TABLE IF NOT EXISTS departed_students (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        student_ref    INTEGER,
        student_name   TEXT,
        form           TEXT,
        reason         TEXT DEFAULT 'Left school',
        departure_date TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- Full case management plan per student — stored as a JSON blob
    -- Matches the official "Student Attendance Case Management Plan" form
    -- Storing as JSON means adding new form fields never requires a schema change
    CREATE TABLE IF NOT EXISTS case_plans (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        student_ref   INTEGER UNIQUE NOT NULL,
        plan_data     TEXT DEFAULT '{}',       -- entire form as JSON
        last_updated  TEXT DEFAULT CURRENT_TIMESTAMP
    );

    -- Day-of-week absence analysis (parsed from Individual Absentee Report)
    CREATE TABLE IF NOT EXISTS day_analysis (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        upload_id INTEGER,
        day_data  TEXT DEFAULT '{}'
    );

    -- Performance indexes — queries filter/join on these columns most frequently
    CREATE INDEX IF NOT EXISTS idx_students_upload_id  ON students(upload_id);
    CREATE INDEX IF NOT EXISTS idx_students_ref        ON students(ref);
    CREATE INDEX IF NOT EXISTS idx_cases_student_ref   ON cases(student_ref);
    CREATE INDEX IF NOT EXISTS idx_case_history_ref    ON case_history(student_ref);
    CREATE INDEX IF NOT EXISTS idx_departed_ref        ON departed_students(student_ref);

    -- Key-value store for app settings (e.g. case plan template)
    CREATE TABLE IF NOT EXISTS settings (
        key   TEXT PRIMARY KEY,
        value TEXT NOT NULL DEFAULT ''
    );

    -- User accounts for authentication
    -- role: 'admin' | 'teacher'
    -- Admin can manage users; teachers access dashboards read/write
    CREATE TABLE IF NOT EXISTS users (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        username      TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        display_name  TEXT NOT NULL,
        role          TEXT NOT NULL DEFAULT 'teacher',  -- 'admin' | 'teacher'
        created_at    TEXT DEFAULT CURRENT_TIMESTAMP,
        created_by    TEXT DEFAULT 'system',
        active        INTEGER DEFAULT 1                  -- 0 = disabled
    );

    """)
    db.commit()

    # Add visible_to_teachers column to uploads if it doesn't exist yet (safe migration)
    existing_cols = [row[1] for row in db.execute("PRAGMA table_info(uploads)").fetchall()]
    if 'visible_to_teachers' not in existing_cols:
        db.execute("ALTER TABLE uploads ADD COLUMN visible_to_teachers INTEGER DEFAULT 1")
        db.commit()
    if 'report_type' not in existing_cols:
        db.execute("ALTER TABLE uploads ADD COLUMN report_type TEXT DEFAULT ''")
        db.commit()

    # Add form_access column to users (safe migration — NULL means all-access)
    user_cols = [row[1] for row in db.execute("PRAGMA table_info(users)").fetchall()]
    if 'form_access' not in user_cols:
        db.execute("ALTER TABLE users ADD COLUMN form_access TEXT DEFAULT NULL")
        db.commit()

    # Add contact_method and contact_outcome to case_history (safe migration)
    ch_cols = [row[1] for row in db.execute("PRAGMA table_info(case_history)").fetchall()]
    if 'contact_method' not in ch_cols:
        db.execute("ALTER TABLE case_history ADD COLUMN contact_method TEXT DEFAULT ''")
        db.commit()
    if 'contact_outcome' not in ch_cols:
        db.execute("ALTER TABLE case_history ADD COLUMN contact_outcome TEXT DEFAULT ''")
        db.commit()

    # Per-upload notes table (notes are specific to each dashboard/report)
    db.execute("""
        CREATE TABLE IF NOT EXISTS upload_notes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            student_ref  INTEGER NOT NULL,
            upload_id    INTEGER NOT NULL,
            notes        TEXT DEFAULT '',
            last_updated TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(student_ref, upload_id)
        )
    """)
    db.commit()

    # Period-based notes (survives upload deletes — keyed on student + period, not upload)
    db.execute("""
        CREATE TABLE IF NOT EXISTS period_notes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            student_ref  INTEGER NOT NULL,
            period_key   TEXT NOT NULL,
            notes        TEXT DEFAULT '',
            last_updated TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(student_ref, period_key)
        )
    """)
    db.commit()

    # Migrate existing upload_notes into period_notes (take latest note per student/period)
    migrated = db.execute("SELECT COUNT(*) FROM period_notes").fetchone()[0]
    if migrated == 0:
        db.execute("""
            INSERT OR IGNORE INTO period_notes (student_ref, period_key, notes, last_updated)
            SELECT
                un.student_ref,
                CASE WHEN LOWER(COALESCE(u.report_type,'')) = 'ytd'
                     THEN 'YTD ' || COALESCE(SUBSTR(u.term, -4), STRFTIME('%Y','now'))
                     ELSE COALESCE(u.term, 'legacy') END AS period_key,
                un.notes,
                un.last_updated
            FROM upload_notes un
            JOIN uploads u ON u.id = un.upload_id
            WHERE un.notes != ''
            ORDER BY un.last_updated DESC
        """)
        db.commit()

    # Migrate case_plans to per-period: add period_key if missing
    cp_cols = [row[1] for row in db.execute("PRAGMA table_info(case_plans)").fetchall()]
    if 'period_key' not in cp_cols:
        db.execute("""
            CREATE TABLE IF NOT EXISTS case_plans_new (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                student_ref  INTEGER NOT NULL,
                period_key   TEXT NOT NULL DEFAULT 'legacy',
                plan_data    TEXT DEFAULT '{}',
                last_updated TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(student_ref, period_key)
            )
        """)
        db.execute("""
            INSERT INTO case_plans_new (student_ref, period_key, plan_data, last_updated)
            SELECT student_ref, 'legacy', plan_data, last_updated FROM case_plans
        """)
        db.execute("DROP TABLE case_plans")
        db.execute("ALTER TABLE case_plans_new RENAME TO case_plans")
        db.commit()

    # Seed a default admin account on first run if no users exist yet.
    # Admin should change this password immediately after first login.
    existing = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing == 0:
        db.execute(
            "INSERT INTO users (username, password_hash, display_name, role, created_by) VALUES (?,?,?,?,?)",
            ('admin', generate_password_hash('admin123'), 'Administrator', 'admin', 'system')
        )
        db.commit()
        print("✅ Default admin account created  →  username: admin  |  password: admin123")
        print("   ⚠️  Please change the default password after first login.")

    db.close()


# Run database setup every time the server starts (safe — uses IF NOT EXISTS)
init_db()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_departed_refs():
    """
    Return a Python set of student ref numbers who have left the school.

    Using a set (not a list) gives O(1) lookup when checking if a student
    should be excluded — important when filtering hundreds of students.
    """
    db = get_db()
    rows = db.execute("SELECT student_ref FROM departed_students").fetchall()
    db.close()
    return {r['student_ref'] for r in rows}


def parse_xls_file(filepath):
    """
    Parse a Maze/school system attendance XLS or XLSX export file into
    a list of student dictionaries.

    Expected column layout (0-indexed) in the export:
      Col 0  — student ref number (or "Form: ACACIA" section headers)
      Col 1  — student name in "Surname, Firstname" format
      Col 3  — year level (e.g. "05", "PR", "TR")
      Col 4  — form/class name
      Col 8  — sessions attended
      Col 9  — total sessions possible
      Col 10 — sessions absent
      Col 11 — attendance percentage (0-100)

    The file has Form header rows like "Form: ACACIA" between student groups.
    We track current_form so students without a form value in col 4 still
    get the correct class assigned.

    Returns:
      tuple: (students, date_from, date_to)
        students  — list of dicts, one per student
        date_from — string e.g. "26 JAN 2026" (or None if not found)
        date_to   — string e.g. "26 MAR 2026" (or None if not found)
      Returns ([], None, None) on any parse error.
    """
    try:
        print(f"📂 Reading file: {filepath}")

        # Choose the correct pandas engine:
        #   xlrd   — handles legacy .xls (no LibreOffice needed)
        #   openpyxl — handles modern .xlsx
        engine = 'openpyxl' if filepath.lower().endswith('.xlsx') else 'xlrd'

        # Load all sheets, then take the first one
        df_dict = pd.read_excel(filepath, sheet_name=None, header=None, engine=engine)
        sheet = list(df_dict.values())[0]
        print(f"📋 Sheet loaded: {sheet.shape[0]} rows x {sheet.shape[1]} cols")

        students = []
        current_form = ''         # updated as we encounter "Form: X" header rows
        date_from, date_to = None, None

        for _, row in sheet.iterrows():
            val0 = str(row[0]).strip() if pd.notna(row[0]) else ''
            val1 = str(row[1]).strip() if pd.notna(row[1]) else ''

            # Detect form section header rows e.g. "Form: ACACIA"
            if val0.startswith('Form:'):
                current_form = val0.replace('Form:', '').strip()

            # Extract the reporting date range from the file header area
            # Handles both "Date Range : 26 JAN 2026 to 26 MAR 2026" (split across cells)
            # and combined single-cell formats
            if val0 == 'Date Range :' or 'Date Range' in val0:
                try:
                    dates = val1.split(' to ')
                    if len(dates) == 2:
                        date_from = dates[0].strip()
                        date_to   = dates[1].strip()
                except:
                    pass
            if not date_from and 'Date Range' in val0 and pd.notna(row[1]):
                try:
                    combined = val0 + ' ' + str(row[1])
                    m = re.search(r'(\d{1,2}\s+\w+\s+\d{4})\s+to\s+(\d{1,2}\s+\w+\s+\d{4})', combined)
                    if m:
                        date_from, date_to = m.group(1), m.group(2)
                except:
                    pass

            # Attempt to parse this row as student data
            # ValueError/TypeError is expected for header/label rows — skip silently
            try:
                ref      = int(float(str(row[0])))
                name     = str(row[1]).strip() if pd.notna(row[1]) else ''
                year     = str(row[3]).strip() if pd.notna(row[3]) else ''
                form     = str(row[4]).strip() if pd.notna(row[4]) else current_form
                attended = float(row[8])  if pd.notna(row[8])  else 0
                sessions = float(row[9])  if pd.notna(row[9])  else 0
                absences = float(row[10]) if pd.notna(row[10]) else 0
                pct      = round(float(row[11]), 2) if pd.notna(row[11]) else 0

                # Cross-validate the file's pct against the raw session counts.
                # The school system's percentage is the authoritative value,
                # but a large discrepancy often signals a misaligned column or
                # corrupted row — log a warning so it can be investigated.
                if sessions > 0:
                    calc_pct = round(attended / sessions * 100, 2)
                    if abs(calc_pct - pct) > 2:
                        print(f"⚠️  Pct mismatch for {name} (ref {ref}): "
                              f"file={pct}%  calculated={calc_pct}%  "
                              f"(attended={int(attended)} sessions={int(sessions)})")
                # Integrity: attended or absences should not exceed total sessions.
                if attended > sessions or absences > sessions:
                    print(f"⚠️  Invalid counts for {name} (ref {ref}): "
                          f"attended={int(attended)} absences={int(absences)} "
                          f"sessions={int(sessions)} — exceeds total")

                # ref > 1000 filters out header rows that parse as small numbers
                if name and name != 'nan' and ref > 1000:
                    students.append({
                        'ref':          ref,
                        'name':         name,
                        'year':         year,
                        'form':         form,
                        'attended':     int(attended),
                        'sessions':     int(sessions),
                        'absences':     int(absences),
                        'pct':          pct,
                        # Convert sessions to school days (2 sessions = 1 day)
                        'days_attended': attended / 2,
                        'days_total':    sessions / 2,
                        'days_absent':   absences / 2,
                    })
            except (ValueError, TypeError):
                pass   # Not a student row — skip

        print(f"✅ Parsed {len(students)} students | Date: {date_from} → {date_to}")
        return students, date_from, date_to

    except Exception as e:
        import traceback
        print(f"❌ Parse error: {e}")
        print(traceback.format_exc())
        return [], None, None


# =============================================================================
# ROUTES — AUTHENTICATION
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        db  = get_db()
        row = db.execute(
            "SELECT * FROM users WHERE username=? AND active=1", (username,)
        ).fetchone()
        db.close()

        if row and check_password_hash(row['password_hash'], password):
            user = User(row['id'], row['username'], row['role'], row['display_name'])
            login_user(user, remember=False)
            from flask import session
            session.permanent = True
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            error = 'Incorrect username or password.'

    return render_template('login.html', error=error)


def get_period_key(upload):
    """
    Derive a stable period key from an upload record.
    - YTD uploads  → 'YTD 2026'  (year extracted from the term field)
    - Term uploads → 'Term 1 2026' (the term field value directly)
    - Legacy / unknown → 'legacy'
    """
    import re as _re
    term        = (upload.get('term') or upload.get('label') or '').strip()
    report_type = (upload.get('report_type') or '').strip().lower()
    if report_type == 'ytd':
        m    = _re.search(r'\b(\d{4})\b', term)
        year = m.group(1) if m else str(datetime.now().year)
        return f'YTD {year}'
    return term if term else 'legacy'


def plan_has_content(plan_data_str):
    """
    Return True only if a case plan record has at least one meaningful field filled in.
    Prevents blank/auto-created records from appearing in the 'Has Case Plan' filter.
    """
    try:
        plan = json.loads(plan_data_str or '{}')
    except Exception:
        return False
    # Text fields that indicate real content ('date' excluded — it's auto-filled on open)
    text_fields = ['goal', 'strengths', 'barriers', 'learning', 'strategies',
                   'casemanager', 'classes', 'success', 'rewards', 'fu-notes',
                   'agency1', 'agency2', 'dob']
    if any(plan.get(f, '').strip() for f in text_fields):
        return True
    # Checkbox fields
    cb_fields = ['sup_curriculum', 'sup_career', 'sup_basicneeds',
                 'sup_mental', 'sup_behaviour', 'sup_social']
    if any(plan.get(f) for f in cb_fields):
        return True
    return False


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    """Log the current user out and redirect to the login page."""
    logout_user()
    return redirect(url_for('login'))


# =============================================================================
# ROUTES — ADMIN USER MANAGEMENT
# =============================================================================

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    db    = get_db()
    users = db.execute("SELECT * FROM users ORDER BY role DESC, created_at ASC").fetchall()
    # Distinct class list from the most recent parsed upload — used to populate dropdowns
    forms = []
    latest = db.execute(
        "SELECT id FROM uploads WHERE parsed=1 ORDER BY upload_date DESC LIMIT 1"
    ).fetchone()
    if latest:
        rows  = db.execute(
            "SELECT DISTINCT form FROM students WHERE upload_id=? AND form != '' ORDER BY form",
            (latest['id'],)
        ).fetchall()
        forms = [r['form'] for r in rows]
    db.close()
    return render_template('admin_users.html', users=users, forms=forms)


@app.route('/admin/users/create', methods=['POST'])
@login_required
@admin_required
def admin_create_user():
    """
    Create a new user account. Validates username uniqueness, password length,
    and role. Optionally restricts the account to a single class via form_access.
    """
    username     = request.form.get('username', '').strip().lower()
    display_name = request.form.get('display_name', '').strip()
    password     = request.form.get('password', '').strip()
    role         = request.form.get('role', 'teacher')

    if not username or not display_name or not password:
        flash('All fields are required.', 'error')
        return redirect(url_for('admin_users'))
    if role not in ('admin', 'teacher'):
        flash('Invalid role.', 'error')
        return redirect(url_for('admin_users'))
    if len(password) < 6:
        flash('Password must be at least 6 characters.', 'error')
        return redirect(url_for('admin_users'))

    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if existing:
        db.close()
        flash(f'Username "{username}" is already taken.', 'error')
        return redirect(url_for('admin_users'))

    # form_access: empty string or absent → NULL (all classes); a class name → restricted
    form_access = request.form.get('form_access', '').strip() or None

    db.execute(
        "INSERT INTO users (username, password_hash, display_name, role, form_access, created_by) VALUES (?,?,?,?,?,?)",
        (username, generate_password_hash(password), display_name, role, form_access, current_user.username)
    )
    db.commit()
    db.close()
    access_label = f' — restricted to {form_access}' if form_access else ''
    flash(f'Account created for {display_name} ({username}){access_label}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    """Set a new password for any user account. Minimum 6 characters enforced."""
    new_password = request.form.get('new_password', '').strip()
    if len(new_password) < 6:
        flash('Password must be at least 6 characters.', 'error')
        return redirect(url_for('admin_users'))

    db  = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not row:
        db.close()
        flash('User not found.', 'error')
        return redirect(url_for('admin_users'))

    db.execute(
        "UPDATE users SET password_hash=? WHERE id=?",
        (generate_password_hash(new_password), user_id)
    )
    db.commit()
    db.close()
    flash(f'Password reset for {row["display_name"]}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_user(user_id):
    """Enable or disable a user account. Admins cannot disable themselves."""
    if user_id == current_user.id:
        flash('You cannot disable your own account.', 'error')
        return redirect(url_for('admin_users'))

    db  = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if row:
        new_state = 0 if row['active'] else 1
        db.execute("UPDATE users SET active=? WHERE id=?", (new_state, user_id))
        db.commit()
        label = 'enabled' if new_state else 'disabled'
        flash(f'Account {label} for {row["display_name"]}.', 'success')
    db.close()
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/update-class', methods=['POST'])
@login_required
@admin_required
def admin_update_class(user_id):
    """Assign or remove a class restriction for a teacher account."""
    form_access = request.form.get('form_access', '').strip() or None
    db  = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not row:
        db.close()
        flash('User not found.', 'error')
        return redirect(url_for('admin_users'))
    db.execute("UPDATE users SET form_access=? WHERE id=?", (form_access, user_id))
    db.commit()
    db.close()
    label = f'restricted to {form_access}' if form_access else 'all classes (unrestricted)'
    flash(f'{row["display_name"]} — class access updated: {label}.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    """Permanently delete a user. Admins cannot delete themselves."""
    if user_id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin_users'))

    db  = get_db()
    row = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if row:
        db.execute("DELETE FROM users WHERE id=?", (user_id,))
        db.commit()
        flash(f'Account deleted: {row["display_name"]}.', 'success')
    db.close()
    return redirect(url_for('admin_users'))


# =============================================================================
# ROUTES — HTML PAGE RENDERING
# =============================================================================

@app.route('/')
@login_required
def index():
    """
    Control Panel — the home/landing page.
    Shows the upload form, list of all uploads, summary stats,
    case status counts, and departed students manager.
    Stats are calculated from the most recently uploaded file.
    """
    db = get_db()

    # Admins see all uploads; teachers only see uploads marked visible
    if current_user.is_admin:
        uploads = db.execute("SELECT * FROM uploads ORDER BY upload_date DESC").fetchall()
    else:
        uploads = db.execute(
            "SELECT * FROM uploads WHERE visible_to_teachers=1 ORDER BY upload_date DESC"
        ).fetchall()

    departed = db.execute("SELECT * FROM departed_students ORDER BY departure_date DESC").fetchall()

    # Base stats on the most recent upload the current user can see
    if current_user.is_admin:
        latest = db.execute(
            "SELECT * FROM uploads WHERE parsed=1 ORDER BY upload_date DESC LIMIT 1"
        ).fetchone()
    else:
        latest = db.execute(
            "SELECT * FROM uploads WHERE parsed=1 AND visible_to_teachers=1 ORDER BY upload_date DESC LIMIT 1"
        ).fetchone()

    stats = {}
    if latest:
        departed_refs = get_departed_refs()
        students = db.execute(
            "SELECT * FROM students WHERE upload_id=?", (latest['id'],)
        ).fetchall()
        active = [s for s in students if s['ref'] not in departed_refs]
        total_sessions = sum(s['sessions'] or 0 for s in active)
        stats = {
            'total':        len(active),
            'zero':         sum(1 for s in active if s['pct'] == 0),
            'below50':      sum(1 for s in active if s['pct'] < 50),
            'below80':      sum(1 for s in active if s['pct'] < 80),
            'below90':      sum(1 for s in active if s['pct'] < 90),
            'avg':          round(sum(s['pct'] for s in active) / len(active), 1) if active else 0,
            'school_pct':   round(sum(s['attended'] for s in active) / total_sessions * 100, 1) if total_sessions > 0 else 0,
            'last_updated': latest['upload_date'],
            'upload_label': latest['label'] or latest['filename'],
        }

    # Count cases by status for the summary pills
    cases       = db.execute("SELECT status, COUNT(*) as cnt FROM cases GROUP BY status").fetchall()
    case_counts = {r['status']: r['cnt'] for r in cases}
    db.close()

    return render_template('index.html', uploads=uploads, stats=stats,
                           case_counts=case_counts, departed=departed)


@app.route('/dashboard/<int:upload_id>')
@login_required
def dashboard(upload_id):
    """
    Render the main attendance dashboard for a specific upload.

    The dashboard has 5 layers (tabs):
      1. Universal Overview   — school-wide stats and charts
      2. Targeted Follow-Up   — 3-tier risk list with action buttons
      3. Case Management      — full caseload with detailed records
      4. Principal Report     — weekly summary for leadership meetings
      5. All Students         — searchable table of all students

    This route:
      1. Loads the upload metadata
      2. Gets all active students for this upload (departed students excluded)
      3. Merges each student's persistent case status + notes from the cases table
      4. Passes the merged data to dashboard.html where JS renders everything

    The JavaScript in dashboard.html also calls /api/cases/all on page load
    as a second guarantee that statuses are always current.
    """
    db = get_db()
    upload = db.execute("SELECT * FROM uploads WHERE id=?", (upload_id,)).fetchone()
    if not upload:
        return "Upload not found", 404

    # Teachers can only view dashboards that admin has made visible
    if not current_user.is_admin and not upload['visible_to_teachers']:
        return redirect(url_for('index'))

    departed_refs   = get_departed_refs()
    students        = db.execute("SELECT * FROM students WHERE upload_id=?", (upload_id,)).fetchall()
    active_students = [dict(s) for s in students if s['ref'] not in departed_refs]

    # Class-based access control — teachers assigned to a specific class see only that class
    form_filter = getattr(current_user, 'form_access', None)
    if form_filter and not current_user.is_admin:
        active_students = [s for s in active_students if s['form'] == form_filter]

    # Merge persistent case data (notes, status) into each student dict
    # This ensures the dashboard always shows the latest case management state
    # even when viewing older uploads
    upload_dict  = dict(upload)
    period_key   = get_period_key(upload_dict)
    upload_dict['period_key'] = period_key

    cases = {r['student_ref']: dict(r) for r in db.execute("SELECT * FROM cases").fetchall()}
    # has_case_plan is true only if a plan with real content exists for this period or legacy
    plan_refs = {
        r['student_ref'] for r in db.execute(
            "SELECT student_ref, plan_data FROM case_plans WHERE period_key=? OR period_key='legacy'",
            (period_key,)
        ).fetchall()
        if plan_has_content(r['plan_data'])
    }
    # Load period-based notes — survives upload deletes and re-uploads
    period_notes = {r['student_ref']: r['notes'] for r in db.execute(
        "SELECT student_ref, notes FROM period_notes WHERE period_key=?", (period_key,)
    ).fetchall()}
    for s in active_students:
        case            = cases.get(s['ref'], {})
        s['status']     = case.get('status', 'pending')
        # Period notes take priority; fall back to legacy global case note
        s['notes']      = period_notes.get(s['ref'], case.get('notes', ''))
        s['has_case_plan'] = s['ref'] in plan_refs

    db.close()
    print(f"📊 Dashboard {upload_id}: serving {len(active_students)} students "
          f"({'class: ' + form_filter if form_filter else 'all classes'})")
    return render_template('dashboard.html', upload=upload_dict, students=active_students,
                           form_filter=form_filter, cp_template=get_cp_template())


@app.route('/compare')
@login_required
def compare():
    """
    Week vs Week comparison page.
    Loads the list of uploads for the dropdowns — actual comparison
    data is fetched via /api/compare/<id1>/<id2> when the user clicks Compare.
    """
    # Teachers cannot access the comparison tool — redirect to home
    if not current_user.is_admin:
        return redirect(url_for('index'))

    db = get_db()
    uploads = db.execute(
        "SELECT * FROM uploads WHERE parsed=1 ORDER BY upload_date ASC"
    ).fetchall()
    db.close()
    return render_template('compare.html', uploads=[dict(u) for u in uploads])


# =============================================================================
# API — FILE UPLOAD & MANAGEMENT
# =============================================================================

@app.route('/api/upload', methods=['POST'])
@login_required
@admin_required
def upload_file():
    """
    Handle XLS/XLSX file upload from the Control Panel upload form.

    Process:
      1. Validate file type (XLS or XLSX only)
      2. Save file to uploads/ folder with timestamp prefix
      3. Parse the file with parse_xls_file()
      4. Create a record in the uploads table
      5. Insert one row per student into the students table
         (departed students are automatically skipped)
      6. Return JSON with upload_id and redirect URL

    Form fields:
      file        — the XLS/XLSX file (multipart/form-data)
      label       — display name e.g. "Term 1 Cumulative 26 Mar"
      term        — e.g. "Term 1 2026"
      week_number — optional integer week number
    """
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    f        = request.files['file']
    label       = request.form.get('label', '')
    week_num    = request.form.get('week_number', None)
    term        = request.form.get('term', 'Term 1 2026')
    report_type = request.form.get('report_type', '')

    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED:
        return jsonify({'error': f'File type {ext} not supported. Use XLS or XLSX'}), 400

    # Timestamp prefix prevents filename collisions from repeated uploads
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(f.filename)}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)

    students, date_from, date_to = parse_xls_file(filepath)

    if not students:
        return jsonify({'error': (
            'No students found in this file. '
            'Make sure it is the correct attendance export from the school system '
            '(XLS/XLSX with student ref numbers in column A). '
            'Check the file is not empty or password-protected.'
        )}), 400

    db = get_db()

    # Create the upload record
    cur = db.execute(
        "INSERT INTO uploads (filename, label, week_number, term, report_type, date_from, date_to, student_count, parsed) VALUES (?,?,?,?,?,?,?,?,1)",
        (filename, label or filename, week_num, term, report_type, date_from, date_to, len(students))
    )
    upload_id = cur.lastrowid

    # Insert student rows — skip departed students
    departed_refs = get_departed_refs()
    for s in students:
        if s['ref'] not in departed_refs:
            db.execute(
                "INSERT INTO students (ref, name, year, form, upload_id, attended, sessions, absences, pct, days_attended, days_total, days_absent) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (s['ref'], s['name'], s['year'], s['form'], upload_id,
                 s['attended'], s['sessions'], s['absences'], s['pct'],
                 s['days_attended'], s['days_total'], s['days_absent'])
            )

    db.commit()
    db.close()
    print(f"✅ Upload {upload_id}: saved {len(students)} students to DB")

    # Delete the raw XLS now that all data is in the DB — no need to keep it
    try:
        os.remove(filepath)
    except OSError:
        pass

    return jsonify({
        'success':         True,
        'upload_id':       upload_id,
        'students_parsed': len(students),
        'date_from':       date_from,
        'date_to':         date_to,
        'redirect':        f'/dashboard/{upload_id}'
    })


@app.route('/api/upload/<int:upload_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_upload(upload_id):
    """
    Delete an upload and all its student attendance rows.

    Important: case notes and statuses in the cases table are NOT deleted.
    They are keyed by student_ref (not upload_id) so they persist and will
    reappear if the same student is in any future upload.
    """
    db = get_db()
    db.execute("DELETE FROM students WHERE upload_id=?", (upload_id,))
    db.execute("DELETE FROM uploads WHERE id=?", (upload_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})


@app.route('/api/upload/<int:upload_id>/visibility', methods=['POST'])
@login_required
@admin_required
def toggle_upload_visibility(upload_id):
    """Toggle whether teachers can see this upload/dashboard."""
    db  = get_db()
    row = db.execute("SELECT visible_to_teachers FROM uploads WHERE id=?", (upload_id,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Not found'}), 404
    new_val = 0 if row['visible_to_teachers'] else 1
    db.execute("UPDATE uploads SET visible_to_teachers=? WHERE id=?", (new_val, upload_id))
    db.commit()
    db.close()
    return jsonify({'success': True, 'visible': bool(new_val)})


@app.route('/api/uploads')
@login_required
def list_uploads():
    """Return all successfully parsed uploads as JSON, newest first."""
    db = get_db()
    uploads = db.execute("SELECT * FROM uploads WHERE parsed=1 ORDER BY upload_date DESC").fetchall()
    db.close()
    return jsonify([dict(u) for u in uploads])


# =============================================================================
# API — CASE MANAGEMENT
# =============================================================================

@app.route('/api/cases/all')
@login_required
def get_all_cases():
    """
    Return ALL case statuses and notes as a dict keyed by student_ref.

    This is called by the dashboard JavaScript on every page load.
    It guarantees that statuses saved in previous sessions are always
    loaded into the STUDENTS array before any rendering happens —
    even if the Jinja2 template merge somehow missed a record.

    Returns: { student_ref_integer: { "status": "...", "notes": "..." }, ... }

    This is a key part of the data persistence guarantee.
    """
    db = get_db()
    cases = db.execute("SELECT student_ref, status, notes FROM cases").fetchall()
    db.close()
    return jsonify({r['student_ref']: {'status': r['status'], 'notes': r['notes']} for r in cases})


@app.route('/api/case/update', methods=['POST'])
@login_required
def update_case():
    """
    Create or update the case record for a student.

    This is called every time a status button is clicked or notes are changed.
    It's designed to be called frequently (debounced to 500ms in the frontend).

    JSON body fields:
      student_ref  — integer (required)
      student_name — string (for history log display)
      form         — class name
      status       — new status string (optional — only status OR notes needed)
      notes        — case notes text (optional)
      updated_by   — staff identifier (defaults to 'Officer')

    Behaviour:
      - Upsert pattern: creates record if new, updates if exists
      - Only updates the provided fields (status and notes are independent)
      - Logs every status change to case_history for audit trail
      - Status changes from pending→anything and anything→pending are both logged
    """
    data            = request.json
    ref             = data.get('student_ref')
    name            = data.get('student_name', '')
    form            = data.get('form', '')
    new_status      = data.get('status')
    notes           = data.get('notes')
    updated_by      = data.get('updated_by', 'Officer')
    contact_method  = data.get('contact_method', '')
    contact_outcome = data.get('contact_outcome', '')

    if not ref:
        return jsonify({'error': 'student_ref required'}), 400

    db = get_db()
    existing   = db.execute("SELECT * FROM cases WHERE student_ref=?", (ref,)).fetchone()
    old_status = existing['status'] if existing else 'pending'

    if existing:
        # Whitelist of columns that callers are permitted to update.
        # Only these names can appear in the SET clause — no f-string column injection possible.
        ALLOWED_CASE_FIELDS = {'status', 'notes'}

        updates, vals = [], []
        if new_status is not None and 'status' in ALLOWED_CASE_FIELDS:
            updates.append("status=?");  vals.append(new_status)
        if notes is not None and 'notes' in ALLOWED_CASE_FIELDS:
            updates.append("notes=?");   vals.append(notes)
        updates.append("last_updated=?"); vals.append(datetime.now().isoformat())
        updates.append("updated_by=?");   vals.append(updated_by)
        vals.append(ref)
        # Safe: SET clause is built only from string literals in ALLOWED_CASE_FIELDS,
        # never from raw user input. All values remain parameterised.
        db.execute("UPDATE cases SET " + ", ".join(updates) + " WHERE student_ref=?", vals)
    else:
        # First case update for this student — create the record
        db.execute(
            "INSERT INTO cases (student_ref, student_name, form, status, notes, updated_by) VALUES (?,?,?,?,?,?)",
            (ref, name, form, new_status or 'pending', notes or '', updated_by)
        )

    # Log status changes for audit trail (notes-only updates don't create history entries)
    if new_status and new_status != old_status:
        db.execute(
            "INSERT INTO case_history (student_ref, student_name, action, old_status, new_status, notes, contact_method, contact_outcome, updated_by) VALUES (?,?,?,?,?,?,?,?,?)",
            (ref, name, 'status_change', old_status, new_status, notes or '', contact_method, contact_outcome, updated_by)
        )

    db.commit()
    db.close()
    return jsonify({'success': True})


@app.route('/api/notes/<int:upload_id>/<int:ref>', methods=['POST'])
@login_required
def save_upload_note(upload_id, ref):
    """
    Save a note for a student, keyed by period (not upload).
    Notes survive upload deletes and re-uploads for the same period.
    """
    notes      = (request.json or {}).get('notes', '')
    db         = get_db()
    upload     = db.execute("SELECT * FROM uploads WHERE id=?", (upload_id,)).fetchone()
    period_key = get_period_key(dict(upload)) if upload else 'legacy'

    db.execute("""
        INSERT INTO period_notes (student_ref, period_key, notes, last_updated)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(student_ref, period_key) DO UPDATE SET
            notes=excluded.notes, last_updated=excluded.last_updated
    """, (ref, period_key, notes, datetime.now().isoformat()))
    db.commit()
    db.close()
    return jsonify({'success': True})


@app.route('/api/case/<int:ref>')
@login_required
def get_case(ref):
    """
    Get the current case record and recent history for a single student.
    Returns the 20 most recent history entries (newest first).
    """
    db = get_db()
    case    = db.execute("SELECT * FROM cases WHERE student_ref=?", (ref,)).fetchone()
    history = db.execute(
        "SELECT * FROM case_history WHERE student_ref=? ORDER BY timestamp DESC", (ref,)
    ).fetchall()
    db.close()
    return jsonify({
        'case':    dict(case) if case else {},
        'history': [dict(h) for h in history]
    })


# =============================================================================
# API — CASE MANAGEMENT PLANS
# =============================================================================

@app.route('/api/caseplan/<int:ref>', methods=['GET'])
@login_required
def get_caseplan(ref):
    """
    Load the case plan for a student for a specific period.
    ?period_key=Term 1 2026  (or 'YTD 2026', 'legacy', etc.)
    Also returns all other periods that have a saved plan, for copy-forward.
    Falls back to 'legacy' plan if no period-specific plan exists.
    """
    period_key = request.args.get('period_key', 'legacy')
    db         = get_db()

    # Load plan for this exact period only — no cross-period fallback
    row = db.execute(
        "SELECT * FROM case_plans WHERE student_ref=? AND period_key=?",
        (ref, period_key)
    ).fetchone()
    # Only treat as a real plan if it has actual content
    if row and not plan_has_content(row['plan_data']):
        row = None

    # Other periods with real content — shown in copy-forward picker
    other_periods = [
        {'period_key': r['period_key'], 'last_updated': r['last_updated']}
        for r in db.execute(
            "SELECT period_key, plan_data, last_updated FROM case_plans WHERE student_ref=? ORDER BY last_updated DESC",
            (ref,)
        ).fetchall()
        if r['period_key'] != period_key and plan_has_content(r['plan_data'])
    ]

    db.close()
    return jsonify({
        'plan':          json.loads(row['plan_data']) if row else None,
        'plan_period':   row['period_key'] if row else None,
        'other_periods': other_periods,
    })


@app.route('/api/caseplan/<int:ref>', methods=['POST'])
@login_required
def save_caseplan(ref):
    """
    Save or update the case plan for a student for a specific period.
    Body: { plan: {...}, period_key: 'Term 1 2026' }
    Uses upsert on (student_ref, period_key).
    """
    data       = request.json
    plan       = data.get('plan', {})
    period_key = data.get('period_key', 'legacy')
    db         = get_db()

    db.execute("""
        INSERT INTO case_plans (student_ref, period_key, plan_data, last_updated)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(student_ref, period_key) DO UPDATE SET
            plan_data=excluded.plan_data, last_updated=excluded.last_updated
    """, (ref, period_key, json.dumps(plan), datetime.now().isoformat()))

    db.commit()
    db.close()
    return jsonify({'success': True})


# =============================================================================
# API — DEPARTED STUDENTS
# =============================================================================

@app.route('/api/depart', methods=['POST'])
@login_required
@admin_required
def mark_departed():
    """
    Mark a student as departed (left the school).
    They disappear from all dashboards immediately.
    Their case notes, history and plans remain in the database permanently.
    Idempotent — marking the same student twice has no effect.
    """
    data   = request.json
    ref    = data.get('student_ref')
    name   = data.get('student_name', '')
    form   = data.get('form', '')
    reason = data.get('reason', 'Left school')

    db       = get_db()
    existing = db.execute("SELECT id FROM departed_students WHERE student_ref=?", (ref,)).fetchone()
    if not existing:
        db.execute(
            "INSERT INTO departed_students (student_ref, student_name, form, reason) VALUES (?,?,?,?)",
            (ref, name, form, reason)
        )
        db.commit()
    db.close()
    return jsonify({'success': True})


@app.route('/api/depart/<int:ref>', methods=['DELETE'])
@login_required
@admin_required
def unmark_departed(ref):
    """
    Restore a departed student to active enrolment.
    They will reappear in all dashboards on the next page load.
    """
    db = get_db()
    db.execute("DELETE FROM departed_students WHERE student_ref=?", (ref,))
    db.commit()
    db.close()
    return jsonify({'success': True})


@app.route('/api/depart/<int:ref>/permanent', methods=['DELETE'])
@login_required
@admin_required
def delete_departed_permanent(ref):
    """
    Permanently delete a departed student record and all case data.
    Removes the student from departed_students, cases, case_history,
    and case_plans. Raw attendance rows in students table are kept
    for historical reporting integrity.
    """
    db = get_db()
    db.execute("DELETE FROM departed_students WHERE student_ref=?", (ref,))
    db.execute("DELETE FROM case_history WHERE student_ref=?", (ref,))
    db.execute("DELETE FROM case_plans WHERE student_ref=?", (ref,))
    db.execute("DELETE FROM cases WHERE student_ref=?", (ref,))
    db.commit()
    db.close()
    return jsonify({'success': True})


# =============================================================================
# API — CASE PLAN TEMPLATE
# =============================================================================

@app.route('/api/admin/school-settings', methods=['GET'])
@login_required
@admin_required
def get_school_settings_api():
    """Return school name and logo URL as JSON (used by the settings panel)."""
    return jsonify(get_school_settings())


@app.route('/api/admin/school-settings', methods=['POST'])
@login_required
@admin_required
def save_school_settings_api():
    """
    Save school name and/or logo from the settings panel (multipart form).
    Accepts name (text) and logo (file upload — PNG, JPG, SVG or WEBP).
    Replaces any previous custom logo file before saving the new one.
    """
    name = request.form.get('school_name', '').strip()
    logo = request.files.get('logo')
    db   = get_db()
    if name:
        db.execute("INSERT INTO settings(key,value) VALUES('school_name',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (name,))
    if logo and logo.filename:
        ext = os.path.splitext(logo.filename)[1].lower()
        if ext in {'.png', '.jpg', '.jpeg', '.svg', '.webp'}:
            # Remove any previous custom logo
            old_ext = db.execute("SELECT value FROM settings WHERE key='school_logo_ext'").fetchone()
            if old_ext and old_ext['value']:
                old_path = os.path.join('static', LOGO_FILENAME + old_ext['value'])
                if os.path.exists(old_path):
                    os.remove(old_path)
            logo.save(os.path.join('static', LOGO_FILENAME + ext))
            db.execute("INSERT INTO settings(key,value) VALUES('school_logo_ext',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (ext,))
    db.commit()
    db.close()
    return jsonify({'success': True, 'settings': get_school_settings()})


@app.route('/api/admin/school-logo/reset', methods=['POST'])
@login_required
@admin_required
def reset_school_logo():
    """Delete the custom school logo file and revert to the default SVG logo."""
    db  = get_db()
    row = db.execute("SELECT value FROM settings WHERE key='school_logo_ext'").fetchone()
    if row and row['value']:
        path = os.path.join('static', LOGO_FILENAME + row['value'])
        if os.path.exists(path):
            os.remove(path)
    db.execute("DELETE FROM settings WHERE key='school_logo_ext'")
    db.commit()
    db.close()
    return jsonify({'success': True})


@app.route('/api/admin/cp-template', methods=['GET'])
@login_required
@admin_required
def get_cp_template_api():
    """Return the current case plan template as JSON (used by the admin editor)."""
    return jsonify(get_cp_template())


@app.route('/api/admin/cp-template', methods=['POST'])
@login_required
@admin_required
def save_cp_template_api():
    """
    Persist an edited case plan template as JSON in the settings table.
    Body: the full template object with sections and fields keys.
    """
    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': 'No data'}), 400
    db = get_db()
    db.execute(
        "INSERT INTO settings(key,value) VALUES('cp_template',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (json.dumps(data),)
    )
    db.commit()
    db.close()
    return jsonify({'success': True})


# =============================================================================
# API — BACKUP
# =============================================================================

@app.route('/api/backup')
@login_required
@admin_required
def download_backup():
    """
    Download a timestamped copy of the SQLite database.
    Copies to a temp file first so the live DB is never locked mid-transfer.
    """
    stamp    = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = os.path.join(tempfile.gettempdir(), f'attendance_backup_{stamp}.db')
    shutil.copy2(DB_PATH, out_path)
    return send_file(out_path, as_attachment=True,
                     download_name=f'Moil_Attendance_Backup_{stamp}.db')


# =============================================================================
# API — COMPARISON & EXPORT
# =============================================================================

@app.route('/api/compare/<int:id1>/<int:id2>')
@login_required
def compare_uploads(id1, id2):
    """
    Compare two uploads to reveal attendance trends per student.

    For each student:
      in both uploads  → pct_before, pct_after, diff, trend (improved/worsened/stable)
      only in id2      → trend = 'new' (new enrolment or re-enrolled)
      only in id1      → trend = 'removed' (may have left or been departed)

    Trend thresholds:
      improved  = attendance rose by more than 2%
      worsened  = attendance fell by more than 2%
      stable    = change within ±2% (normal weekly variation)

    Departed students are excluded from both sides of the comparison.
    Results sorted lowest attendance first (most urgent at top).
    """
    db            = get_db()
    departed_refs = get_departed_refs()

    def get_students(uid):
        """Load one upload's students as a ref→dict map for fast comparison."""
        rows = db.execute("SELECT * FROM students WHERE upload_id=?", (uid,)).fetchall()
        return {r['ref']: dict(r) for r in rows if r['ref'] not in departed_refs}

    s1 = get_students(id1)   # "before" upload
    s2 = get_students(id2)   # "after" upload
    u1 = dict(db.execute("SELECT * FROM uploads WHERE id=?", (id1,)).fetchone())
    u2 = dict(db.execute("SELECT * FROM uploads WHERE id=?", (id2,)).fetchone())
    db.close()

    comparison = []
    for ref in set(s1.keys()) | set(s2.keys()):
        a = s1.get(ref)
        b = s2.get(ref)
        if a and b:
            diff  = round(b['pct'] - a['pct'], 2)
            trend = 'improved' if diff > 2 else 'worsened' if diff < -2 else 'stable'
            comparison.append({
                'ref': ref, 'name': b['name'], 'form': b['form'], 'year': b['year'],
                'pct_before': a['pct'], 'pct_after': b['pct'],
                'diff': diff, 'trend': trend,
                'abs_before': a['absences'], 'abs_after': b['absences'],
            })
        elif b:
            comparison.append({'ref': ref, 'name': b['name'], 'form': b['form'],
                               'pct_before': None, 'pct_after': b['pct'], 'trend': 'new'})
        elif a:
            comparison.append({'ref': ref, 'name': a['name'], 'form': a['form'],
                               'pct_before': a['pct'], 'pct_after': None, 'trend': 'removed'})

    comparison.sort(key=lambda x: (x.get('pct_after') or 0))
    return jsonify({'upload1': u1, 'upload2': u2, 'students': comparison})


@app.route('/api/export/<int:upload_id>')
@login_required
def export_csv(upload_id):
    """Export attendance + case data for a specific upload as a formatted Excel file."""
    db            = get_db()
    departed_refs = get_departed_refs()
    students      = db.execute("SELECT * FROM students WHERE upload_id=?", (upload_id,)).fetchall()
    cases         = {r['student_ref']: dict(r) for r in db.execute("SELECT * FROM cases").fetchall()}
    upload_notes  = {r['student_ref']: r['notes'] for r in db.execute(
        "SELECT student_ref, notes FROM upload_notes WHERE upload_id=?", (upload_id,)).fetchall()}
    upload        = db.execute("SELECT * FROM uploads WHERE id=?", (upload_id,)).fetchone()
    db.close()

    label      = upload['label'] if upload else f'Upload {upload_id}'
    safe_label = re.sub(r'[^\w\-]', '_', label)
    term       = upload['term'] if upload else ''
    date_range = f"{upload['date_from']} – {upload['date_to']}" if upload and upload['date_from'] else ''

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'

    # ── colour palette ──
    green_dark  = PatternFill('solid', fgColor='166534')
    green_light = PatternFill('solid', fgColor='F0FDF4')
    amber_fill  = PatternFill('solid', fgColor='FFFBEB')
    red_fill    = PatternFill('solid', fgColor='FEF2F2')
    blue_fill   = PatternFill('solid', fgColor='EFF6FF')
    grey_fill   = PatternFill('solid', fgColor='F8FAFC')
    hdr_fill    = PatternFill('solid', fgColor='1E293B')

    white_bold   = Font(bold=True, color='FFFFFF', size=11)
    white_normal = Font(color='FFFFFF', size=10)
    dark_bold    = Font(bold=True, color='1E293B', size=10)
    dark_normal  = Font(color='1E293B', size=10)
    thin_border  = Border(
        bottom=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin',  color='E2E8F0'),
    )

    # ── school header (rows 1-3) ──
    ws.merge_cells('A1:K1')
    ws['A1'] = 'MOIL PRIMARY SCHOOL — Attendance Report'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = green_dark
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    ws['A2'] = f'{label}   |   {term}   |   {date_range}   |   Exported {datetime.now().strftime("%d %b %Y")}'
    ws['A2'].font = Font(color='FFFFFF', size=10)
    ws['A2'].fill = PatternFill('solid', fgColor='166534')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6  # spacer

    # ── column headers (row 4) ──
    headers = ['#', 'Ref', 'Student Name', 'Form', 'Year',
               'Days Attended', 'School Days', 'Days Absent', 'Attendance %',
               'Risk Level', 'Case Status', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font    = white_bold
        cell.fill    = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border  = thin_border
    ws.row_dimensions[4].height = 22

    # ── data rows ──
    STATUS_LABELS = {
        'pending':'Pending','contacted':'Contacted','meeting':'Meeting Arranged',
        'welfare':'Welfare Referral','referred':'Principal Referral',
        'agency':'Multi-Agency','resolved':'Resolved','watchlist':'Watchlist'
    }
    row_num = 5
    for s in sorted(students, key=lambda x: x['pct']):
        if s['ref'] in departed_refs:
            continue
        case   = cases.get(s['ref'], {})
        status = case.get('status', 'pending')
        notes  = upload_notes.get(s['ref'], case.get('notes', ''))
        pct    = s['pct']
        days_a = round(s['attended'] / 2, 1) if s.get('attended') is not None else round(s.get('days_attended', 0) / 2, 1)
        days_t = round(s['sessions'] / 2, 1) if s.get('sessions') is not None else round(s.get('days_total', 0) / 2, 1)
        days_ab= round(s['absences'] / 2, 1) if s.get('absences') is not None else round(s.get('days_absent', 0) / 2, 1)

        if pct == 0:   risk, row_fill = 'Zero',     red_fill
        elif pct < 50: risk, row_fill = 'Critical', red_fill
        elif pct < 80: risk, row_fill = 'Concern',  amber_fill
        elif pct < 90: risk, row_fill = 'Watch',    blue_fill
        else:          risk, row_fill = 'Good',      green_light

        values = [row_num - 4, s['ref'], s['name'], s['form'], s['year'],
                  days_a, days_t, days_ab, f"{pct}%",
                  risk, STATUS_LABELS.get(status, status.title()), notes]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill   = row_fill if col > 1 else grey_fill
            cell.font   = dark_bold if col in (3, 9) else dark_normal
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical='center',
                horizontal='center' if col not in (3, 12) else 'left',
                wrap_text=(col == 12)
            )
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # ── summary row ──
    ws.row_dimensions[row_num].height = 6
    row_num += 1
    ws.merge_cells(f'A{row_num}:K{row_num}')
    ws[f'A{row_num}'] = f'Total: {row_num - 6} students exported   |   Generated by Moil Primary School Attendance Dashboard'
    ws[f'A{row_num}'].font = Font(color='94A3B8', size=9, italic=True)
    ws[f'A{row_num}'].alignment = Alignment(horizontal='center')

    # ── column widths ──
    col_widths = [5, 10, 28, 16, 8, 14, 13, 13, 14, 12, 18, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A5'

    out_path = os.path.join(tempfile.gettempdir(), f'export_{safe_label}.xlsx')
    wb.save(out_path)
    return send_file(out_path, as_attachment=True,
                     download_name=f'Moil_Attendance_{safe_label}.xlsx')


@app.route('/api/export/student/<int:ref>')
@login_required
def export_student_csv(ref):
    """Export full case history for a single student as a formatted Excel file."""
    db      = get_db()
    student = db.execute(
        """SELECT s.* FROM students s JOIN uploads u ON s.upload_id=u.id
           WHERE s.ref=? AND u.parsed=1 ORDER BY u.upload_date DESC LIMIT 1""",
        (ref,)
    ).fetchone()
    case    = db.execute("SELECT * FROM cases WHERE student_ref=?", (ref,)).fetchone()
    history = db.execute(
        "SELECT * FROM case_history WHERE student_ref=? ORDER BY timestamp ASC", (ref,)
    ).fetchall()
    db.close()

    name      = student['name'] if student else f'Student_{ref}'
    form      = student['form'] if student else ''
    year      = student['year'] if student else ''
    pct       = student['pct']  if student else 0
    status    = case['status']  if case    else 'pending'
    notes     = case['notes']   if case    else ''
    safe_name = re.sub(r'[^\w\-]', '_', name)

    STATUS_LABELS = {
        'pending':'Pending','contacted':'Contacted','meeting':'Meeting Arranged',
        'welfare':'Welfare Referral','referred':'Principal Referral',
        'agency':'Multi-Agency','resolved':'Resolved','watchlist':'Watchlist'
    }

    wb = Workbook()
    ws = wb.active
    ws.title = 'Student Case Report'

    # ── colour palette ──
    green_dark  = PatternFill('solid', fgColor='166534')
    hdr_fill    = PatternFill('solid', fgColor='1E293B')
    info_fill   = PatternFill('solid', fgColor='F8FAFC')
    amber_fill  = PatternFill('solid', fgColor='FFFBEB')
    red_fill    = PatternFill('solid', fgColor='FEF2F2')
    blue_fill   = PatternFill('solid', fgColor='EFF6FF')
    green_light = PatternFill('solid', fgColor='F0FDF4')
    purple_fill = PatternFill('solid', fgColor='F5F3FF')
    section_fill= PatternFill('solid', fgColor='334155')

    white_bold   = Font(bold=True, color='FFFFFF', size=11)
    dark_bold    = Font(bold=True, color='1E293B', size=10)
    dark_normal  = Font(color='1E293B', size=10)
    label_font   = Font(bold=True, color='475569', size=9)
    thin_border  = Border(
        bottom=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin',  color='E2E8F0'),
    )
    center = Alignment(horizontal='center', vertical='center')

    NUM_COLS = 8

    def merge_header(row, text, fill, font, height=26):
        ws.merge_cells(f'A{row}:{get_column_letter(NUM_COLS)}{row}')
        c = ws[f'A{row}']
        c.value = text; c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = height

    # ── row 1: school banner ──
    merge_header(1, 'MOIL PRIMARY SCHOOL — Student Case Report',
                 green_dark, Font(bold=True, color='FFFFFF', size=14), 30)

    # ── row 2: export date ──
    merge_header(2, f'Exported {datetime.now().strftime("%d %b %Y")}',
                 PatternFill('solid', fgColor='166534'),
                 Font(color='FFFFFF', size=10), 18)

    # ── row 3: spacer ──
    ws.row_dimensions[3].height = 8

    # ── rows 4-9: student info block ──
    info_pairs = [
        ('Name',               name),
        ('Student Ref',        str(ref)),
        ('Form / Class',       form),
        ('Year Group',         str(year)),
        ('Current Attendance', f'{pct}%'),
        ('Current Status',     STATUS_LABELS.get(status, status.title())),
    ]
    # determine row background for attendance
    if pct == 0:        att_fill = red_fill
    elif pct < 50:      att_fill = red_fill
    elif pct < 80:      att_fill = amber_fill
    elif pct < 90:      att_fill = blue_fill
    else:               att_fill = green_light

    for i, (lbl, val) in enumerate(info_pairs, 4):
        fill = att_fill if lbl == 'Current Attendance' else info_fill
        # label cell (cols A-B merged)
        ws.merge_cells(f'A{i}:B{i}')
        lc = ws[f'A{i}']
        lc.value = lbl; lc.font = label_font; lc.fill = fill
        lc.alignment = Alignment(horizontal='right', vertical='center')
        lc.border = thin_border
        # value cell (cols C-H merged)
        ws.merge_cells(f'C{i}:{get_column_letter(NUM_COLS)}{i}')
        vc = ws[f'C{i}']
        vc.value = val
        vc.font  = Font(bold=True, color='1E293B', size=11) if lbl in ('Name', 'Current Attendance') else dark_normal
        vc.fill  = fill
        vc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        vc.border = thin_border
        ws.row_dimensions[i].height = 20

    # ── notes row ──
    note_row = 10
    ws.merge_cells(f'A{note_row}:B{note_row}')
    lc = ws[f'A{note_row}']
    lc.value = 'Notes'; lc.font = label_font; lc.fill = info_fill
    lc.alignment = Alignment(horizontal='right', vertical='center')
    lc.border = thin_border
    ws.merge_cells(f'C{note_row}:{get_column_letter(NUM_COLS)}{note_row}')
    vc = ws[f'C{note_row}']
    vc.value = notes or ''; vc.font = dark_normal; vc.fill = info_fill
    vc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    vc.border = thin_border
    ws.row_dimensions[note_row].height = 36 if notes else 20

    # ── spacer ──
    spacer1 = note_row + 1
    ws.row_dimensions[spacer1].height = 8

    # ── section header: Contact History ──
    sec_row = spacer1 + 1
    merge_header(sec_row, 'CONTACT HISTORY',
                 section_fill, Font(bold=True, color='FFFFFF', size=11), 22)

    # ── column headers ──
    col_headers = ['Date', 'Time', 'Previous Status', 'New Status',
                   'Contact Method', 'Outcome', 'Notes', 'Updated By']
    hdr_row = sec_row + 1
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF', size=9)
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[hdr_row].height = 20
    ws.freeze_panes = f'A{hdr_row + 1}'

    # ── status colour map ──
    STATUS_FILL = {
        'pending':   PatternFill('solid', fgColor='F8FAFC'),
        'contacted': blue_fill,
        'meeting':   PatternFill('solid', fgColor='ECFDF5'),
        'welfare':   amber_fill,
        'referred':  PatternFill('solid', fgColor='FEF3C7'),
        'agency':    red_fill,
        'resolved':  green_light,
        'watchlist': purple_fill,
    }

    data_row = hdr_row + 1
    for h in history:
        ts      = h['timestamp'] or ''
        date    = ts[:10] if len(ts) >= 10 else ts
        time_s  = ts[11:16] if len(ts) >= 16 else ''
        method  = (h['contact_method']  or '') if 'contact_method'  in h.keys() else ''
        outcome = (h['contact_outcome'] or '') if 'contact_outcome' in h.keys() else ''
        h_notes = (h['notes'] or '').replace('\n', ' ')
        new_st  = h['new_status'] or ''
        row_fill = STATUS_FILL.get(new_st, PatternFill('solid', fgColor='F8FAFC'))

        values = [date, time_s,
                  STATUS_LABELS.get(h['old_status'], (h['old_status'] or '').title()),
                  STATUS_LABELS.get(new_st, new_st.title()),
                  method, outcome, h_notes, h['updated_by'] or '']

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.fill      = row_fill
            cell.font      = dark_bold if col == 4 else dark_normal
            cell.border    = thin_border
            cell.alignment = Alignment(
                horizontal='left' if col in (5, 6, 7) else 'center',
                vertical='center', wrap_text=(col == 7)
            )
        ws.row_dimensions[data_row].height = 30 if h_notes else 18
        data_row += 1

    if data_row == hdr_row + 1:
        ws.merge_cells(f'A{data_row}:{get_column_letter(NUM_COLS)}{data_row}')
        c = ws[f'A{data_row}']
        c.value = 'No contact history recorded.'
        c.font = Font(color='94A3B8', italic=True, size=10)
        c.alignment = center
        ws.row_dimensions[data_row].height = 20
        data_row += 1

    # ── footer ──
    footer_row = data_row + 1
    ws.merge_cells(f'A{footer_row}:{get_column_letter(NUM_COLS)}{footer_row}')
    fc = ws[f'A{footer_row}']
    fc.value = f'Generated by Moil Primary School Attendance Dashboard   |   {datetime.now().strftime("%d %b %Y %H:%M")}'
    fc.font = Font(color='94A3B8', size=9, italic=True)
    fc.alignment = center

    # ── column widths ──
    col_widths = [12, 8, 20, 20, 20, 20, 38, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out_path = os.path.join(tempfile.gettempdir(), f'case_{safe_name}_{ref}.xlsx')
    wb.save(out_path)
    return send_file(out_path, as_attachment=True,
                     download_name=f'Case_{safe_name}.xlsx')


@app.route('/api/students/latest')
@login_required
def latest_students():
    """
    Return students from the most recent upload with case statuses merged.
    Fallback endpoint for getting current data without knowing the upload ID.
    """
    db            = get_db()
    departed_refs = get_departed_refs()
    latest        = db.execute(
        "SELECT * FROM uploads WHERE parsed=1 ORDER BY upload_date DESC LIMIT 1"
    ).fetchone()

    if not latest:
        db.close()
        return jsonify({'students': [], 'upload': {}})

    students = db.execute("SELECT * FROM students WHERE upload_id=?", (latest['id'],)).fetchall()
    cases    = {r['student_ref']: dict(r) for r in db.execute("SELECT * FROM cases").fetchall()}
    db.close()

    result = []
    for s in students:
        if s['ref'] not in departed_refs:
            sd           = dict(s)
            case         = cases.get(s['ref'], {})
            sd['status'] = case.get('status', 'pending')
            sd['notes']  = case.get('notes', '')
            result.append(sd)

    return jsonify({'students': result, 'upload': dict(latest)})


# =============================================================================
# API — STUDENT TREND (Weekly History)
# =============================================================================

@app.route('/api/trend/<int:ref>')
@login_required
def student_trend(ref):
    """
    Return a student's attendance history across ALL uploads — ordered by date.
    Used to draw the weekly trend graph on each student card.

    Returns a list of data points, one per upload the student appears in:
      { label, upload_date, pct, attended, absences, sessions }

    This lets the dashboard show whether attendance is improving,
    declining or staying flat across weeks/uploads.
    """
    db = get_db()
    # Join students with uploads to get label + date alongside attendance data
    rows = db.execute("""
        SELECT s.pct, s.attended, s.absences, s.sessions,
               u.label, u.upload_date, u.date_from, u.date_to
        FROM students s
        JOIN uploads u ON s.upload_id = u.id
        WHERE s.ref = ? AND u.parsed = 1
        ORDER BY u.upload_date ASC
    """, (ref,)).fetchall()
    db.close()

    trend = []
    for r in rows:
        trend.append({
            'label':       r['label'] or r['upload_date'][:10],
            'upload_date': r['upload_date'][:10],
            'date_from':   r['date_from'],
            'date_to':     r['date_to'],
            'pct':         r['pct'],
            'attended':    r['attended'],
            'absences':    r['absences'],
            'sessions':    r['sessions'],
            'days_absent': round(r['absences'] / 2, 1),
        })

    return jsonify({'ref': ref, 'trend': trend})


def parse_absentee_file(filepath):
    """
    Parse the Individual Absentee Report XLS to extract day-of-week absence patterns.
    Returns a dict: { student_name: { Mon:N, Tue:N, Wed:N, Thu:N, Fri:N, form:X, total:N } }

    Absence codes counted as absent: U (Unnotified), N (Sanctioned), X (Unacceptable),
    H (Internal Suspension), Z (Suspended), K (Community Unrest), S (Notified sick)
    """
    try:
        print(f"📊 Parsing absentee file: {os.path.getsize(filepath)} bytes")
        engine = 'openpyxl' if filepath.lower().endswith('.xlsx') else 'xlrd'
        print(f"📊 Using engine: {engine}")
        df_dict = pd.read_excel(filepath, sheet_name=None, header=None, engine=engine)
        sheet = list(df_dict.values())[0]
        print(f"📊 Sheet loaded: {sheet.shape}")

        ABSENT_CODES = {'U', 'N', 'X', 'H', 'Z', 'K', 'S'}
        students_data = {}
        current_name = None
        in_absences = False

        for _, row in sheet.iterrows():
            val0 = str(row[0]).strip() if pd.notna(row[0]) else ''
            val6 = str(row[6]).strip() if pd.notna(row[6]) else ''

            # Detect student name row
            if 'Form:' in val6 and val0 and val0 not in ['Moil Primary School', 'Individual Absentee Report', 'Date Range :']:
                current_name = val0.strip()
                form = val6.replace('Form:', '').strip()
                in_absences = False
                if current_name not in students_data:
                    students_data[current_name] = {'form': form, 'Mon': 0, 'Tue': 0, 'Wed': 0, 'Thu': 0, 'Fri': 0, 'total': 0}

            if val0 == 'Day/Date':
                in_absences = True
                continue

            if 'Total Half-Days' in val0 or 'Attendance Codes:' in val0:
                in_absences = False
                continue

            if in_absences and current_name and val0:
                date_match = re.match(r'(\w+),\s+\d+\w*\s+\w+', val0)
                if date_match:
                    day = date_match.group(1)
                    am = str(row[2]).strip() if pd.notna(row[2]) else ''
                    pm = str(row[4]).strip() if pd.notna(row[4]) else ''
                    if (am in ABSENT_CODES or pm in ABSENT_CODES) and day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                        students_data[current_name][day] += 1
                        students_data[current_name]['total'] += 1

        print(f"✅ Absentee report: {len(students_data)} students parsed")
        return students_data
    except Exception as e:
        import traceback
        print(f"❌ Absentee parse error: {e}")
        print(traceback.format_exc())
        return {}


@app.route('/api/upload/absentee', methods=['POST'])
@login_required
@admin_required
def upload_absentee():
    """
    Upload and parse an Individual Absentee Report XLS.

    Two modes:
      term   — replaces existing day analysis with full term data
      week   — merges new week's data into existing analysis (adds counts)

    Weekly merging: if existing_data is sent in the request, the new
    week's absence counts are ADDED to the existing counts per student.
    This builds up a full picture over time from weekly uploads.
    """
    print(f"📥 /api/upload/absentee called, files: {list(request.files.keys())}, form: {list(request.form.keys())}")
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    f         = request.files['file']
    print(f"📥 File received: {f.filename}, size will be saved to disk")
    upload_id = request.form.get('upload_id')
    period    = request.form.get('period', 'term')  # 'term' or 'week'
    existing_json = request.form.get('existing_data', None)

    if not f.filename:
        return jsonify({'error': 'No file selected'}), 400

    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(f.filename)}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)

    new_data = parse_absentee_file(filepath)
    if not new_data:
        return jsonify({'error': 'Could not parse absentee file'}), 400

    # Weekly mode: merge new week into existing data
    if period == 'week' and existing_json:
        try:
            existing = json.loads(existing_json)
            for name, new_counts in new_data.items():
                if name in existing:
                    # Add new week's counts to existing counts
                    for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']:
                        existing[name][day] = existing[name].get(day, 0) + new_counts.get(day, 0)
                    existing[name]['total'] = existing[name].get('total', 0) + new_counts.get('total', 0)
                else:
                    existing[name] = new_counts
            merged_data = existing
            print(f"✅ Weekly merge: {len(new_data)} new + existing = {len(merged_data)} total students")
        except Exception as e:
            print(f"Merge error: {e} — using new data only")
            merged_data = new_data
    else:
        merged_data = new_data

    # Save merged/new data to database
    if upload_id:
        db = get_db()
        existing_row = db.execute(
            "SELECT id FROM day_analysis WHERE upload_id=?", (upload_id,)
        ).fetchone()
        if existing_row:
            db.execute(
                "UPDATE day_analysis SET day_data=? WHERE upload_id=?",
                (json.dumps(merged_data), upload_id)
            )
        else:
            db.execute(
                "INSERT INTO day_analysis (upload_id, day_data) VALUES (?,?)",
                (upload_id, json.dumps(merged_data))
            )
        db.commit()
        db.close()

    day_totals = {d: sum(s.get(d, 0) for s in merged_data.values()) for d in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']}
    worst_day  = max(day_totals, key=day_totals.get)

    return jsonify({
        'success':   True,
        'students':  len(merged_data),
        'period':    period,
        'day_totals': day_totals,
        'worst_day': worst_day,
        'data':      merged_data
    })



def get_dayofweek(upload_id):
    """Return stored day-of-week absence data for this upload."""
    db = get_db()
    row = db.execute(
        "SELECT day_data FROM day_analysis WHERE upload_id=?", (upload_id,)
    ).fetchone()
    db.close()
    if row:
        return jsonify({'data': json.loads(row['day_data'])})
    return jsonify({'data': None})


@app.route('/api/dayofweek/<int:upload_id>', methods=['POST'])
@login_required
@admin_required
def save_dayofweek(upload_id):
    """Save parsed day-of-week absence data for an upload."""
    data = request.json.get('data', {})
    db = get_db()
    existing = db.execute(
        "SELECT id FROM day_analysis WHERE upload_id=?", (upload_id,)
    ).fetchone()
    if existing:
        db.execute(
            "UPDATE day_analysis SET day_data=? WHERE upload_id=?",
            (json.dumps(data), upload_id)
        )
    else:
        db.execute(
            "INSERT INTO day_analysis (upload_id, day_data) VALUES (?,?)",
            (upload_id, json.dumps(data))
        )
    db.commit()
    db.close()
    return jsonify({'success': True})


# =============================================================================
# ROUTES — DAY OF WEEK ANALYSIS (page renderer + form-POST upload)
# =============================================================================


@app.route('/dayanalysis/<int:upload_id>')
@login_required
def dayanalysis_page(upload_id):
    """
    Render the standalone Day of Week Analysis page (loaded inside an iframe
    on the dashboard). Reads stored analysis data and builds a self-contained
    HTML page with a bar chart, pattern insights, and a filterable student table.
    Non-admin users see only their own class.
    """
    db = get_db()
    upload = db.execute("SELECT * FROM uploads WHERE id=?", (upload_id,)).fetchone()
    existing = db.execute("SELECT day_data FROM day_analysis WHERE upload_id=?", (upload_id,)).fetchone()
    db.close()
    data = json.loads(existing['day_data']) if existing else None
    # Filter to own class for non-admin users
    fa = getattr(current_user, 'form_access', None)
    if data and fa:
        data = {name: s for name, s in data.items() if s.get('form') == fa}
    days = ['Mon','Tue','Wed','Thu','Fri']
    day_colors = {'Mon':'#1A4F7A','Tue':'#2E7D32','Wed':'#6B2FAA','Thu':'#D35400','Fri':'#C0392B'}
    patterns = {}
    day_totals = {}
    if data:
        day_totals = {d: sum(s.get(d,0) for s in data.values()) for d in days}
        worst = max(day_totals, key=day_totals.get)
        best_day = min(day_totals, key=day_totals.get)
        max_count = max(day_totals.values()) or 1
        for name, s in data.items():
            vals = {d: s.get(d,0) for d in days}
            total = s.get('total', sum(vals.values()))
            max_day = max(vals, key=vals.get)
            max_val = vals[max_day]
            fri = vals['Fri']
            mon = vals['Mon']
            fri_mon = fri + mon
            mid = vals['Tue'] + vals['Wed'] + vals['Thu']
            if max_val == 0: pat = 'none'
            elif fri >= 5 and fri == max_val: pat = 'friday_always'
            elif fri >= 3 and fri >= mon and fri >= vals['Wed']: pat = 'friday_often'
            elif mon >= 3 and mon == max_val: pat = 'monday_often'
            elif fri_mon >= 5 and fri_mon > mid: pat = 'weekend_extended'
            elif vals['Wed'] >= 3 and vals['Wed'] == max_val: pat = 'midweek'
            elif total >= 10 and max_val <= total // 5 + 1: pat = 'random'
            else: pat = 'spread'
            patterns[name] = {'pattern': pat, 'max_day': max_day, 'max_val': max_val,
                'fri': fri, 'mon': mon, 'total': total, 'form': s.get('form','--'), 'days': vals}

    # Build upload form
    upload_form = """<div class="upload-form"><details><summary style="cursor:pointer;font-weight:700;color:#1A5C1A;">+ Upload New File (click to expand)</summary><div style="margin-top:12px;"><form method="POST" action="/dayanalysis/{uid}" enctype="multipart/form-data"><div style="display:flex;gap:8px;margin-bottom:10px;"><div class="modebt on" id="bt" onclick="sel('term')">Full Term<div style="font-size:11px;color:#888;">Replaces previous</div></div><div class="modebt" id="bw" onclick="sel('week')">Weekly<div style="font-size:11px;color:#888;">Adds to existing</div></div></div><input type="hidden" name="period" id="pi" value="term"><div class="file-input"><input type="file" name="file" accept=".xls,.xlsx" required></div><button type="submit" class="sbtn">Analyse</button></form></div></details></div>""".format(uid=upload_id)

    if not data:
        # No data yet — show upload form prominently
        content = "<div class=\'topbar\'><a href=\'/dashboard/{uid}\' target=\'_top\'>Back to Dashboard</a><h2>Day of Week Analysis</h2></div><div class=\'content\'>{form}</div>".format(uid=upload_id, form=upload_form)
        content = upload_form
        return "<!DOCTYPE html><html><head><meta charset=\'UTF-8\'><title>Day Analysis</title></head><body>" + content + "</body></html>"

    # Has data - render full results
    worst = max(day_totals, key=day_totals.get)
    best_day = min(day_totals, key=day_totals.get)
    max_count = max(day_totals.values()) or 1

    # Bars
    bars = ""
    for d in days:
        pct = int(day_totals[d] / max_count * 100)
        clr = '#C0392B' if d == worst else day_colors[d]
        bld = 'bold' if d == worst else 'normal'
        bars += '<div class="bar-row"><span class="bar-label" style="color:' + clr + ';font-weight:' + bld + ';">' + d + '</span><div class="bar-track"><div class="bar-fill" style="background:' + clr + ';width:' + str(pct) + '%;">' + str(day_totals[d]) + '</div></div><span class="bar-count" style="color:' + clr + ';font-weight:' + bld + ';">' + str(day_totals[d]) + '</span></div>'

    # Pattern counts
    pats = {}
    for p in patterns.values():
        pats[p['pattern']] = pats.get(p['pattern'],0) + 1

    fri_n = pats.get('friday_always',0) + pats.get('friday_often',0)
    mon_n = pats.get('monday_often',0) + pats.get('weekend_extended',0)
    med_n = pats.get('random',0)

    insights = ""
    if fri_n > 0:
        insights += '<div class="pat-card" style="background:#FFEBEE;border-color:#C0392B;"><strong style="color:#C0392B">Friday Pattern — ' + str(fri_n) + ' students</strong><p>These students consistently miss Fridays — likely extending the weekend. Recommend: Friday morning engagement calls, attendance reward programs, family conversation about weekend activities planning.</p></div>'
    if mon_n > 0:
        insights += '<div class="pat-card" style="background:#FFF3E0;border-color:#D35400;"><strong style="color:#D35400">Monday/Weekend Pattern — ' + str(mon_n) + ' students</strong><p>Missing Mondays or Fri+Mon suggests weekend activities running long or difficulty transitioning. Recommend: Monday check-in programs, Sunday family engagement.</p></div>'
    if med_n > 0:
        insights += '<div class="pat-card" style="background:#E8F5E9;border-color:#2E7D32;"><strong style="color:#2E7D32">Spread/Medical Pattern — ' + str(med_n) + ' students</strong><p>Absences spread across all days — suggests illness, chronic conditions, or family circumstances rather than avoidance. Consider medical review or health/NDIS referral.</p></div>'
    insights += '<div class="pat-card" style="background:#E3F2FD;border-color:#1A4F7A;"><strong style="color:#1A4F7A">School-Wide: ' + worst + ' is the hardest day</strong><p>' + worst + ' has ' + str(day_totals[worst]) + ' absent sessions vs ' + best_day + ' (' + str(day_totals[best_day]) + ') — a difference of ' + str(day_totals[worst]-day_totals[best_day]) + ' sessions. What makes ' + best_day + ' more engaging?</p></div>'

    # Table rows
    pat_info = {
        'friday_always': ('Always misses Fri', '#FFEBEE', '#C0392B'),
        'friday_often': ('Often misses Fri', '#FFF3E0', '#D35400'),
        'monday_often': ('Often misses Mon', '#FFF3E0', '#D35400'),
        'weekend_extended': ('Extended weekend', '#FFF3E0', '#E65100'),
        'midweek': ('Midweek absences', '#F3E5F5', '#6B2FAA'),
        'random': ('Spread/Medical', '#E8F5E9', '#2E7D32'),
        'spread': ('General absences', '#F5F5F5', '#666'),
        'none': ('No pattern', '#F5F5F5', '#999'),
    }
    tbody = ""
    forms_set = set()
    for name, p in sorted(patterns.items(), key=lambda x: x[1]['total'], reverse=True):
        if p['total'] < 1: continue
        forms_set.add(p['form'])
        pl, pbg, pclr = pat_info.get(p['pattern'], ('?','#eee','#333'))
        dcells = ""
        for d in days:
            n = p['days'].get(d,0)
            clr = day_colors[d]
            if n >= 5: cs = 'background:' + clr + ';color:white;font-weight:bold;border-radius:4px;padding:2px 5px;display:inline-block;'
            elif n >= 3: cs = 'background:' + clr + '33;color:' + clr + ';font-weight:bold;border-radius:4px;padding:2px 5px;display:inline-block;'
            elif n >= 1: cs = 'color:' + clr + ';font-weight:600;'
            else: cs = 'color:#ccc;'
            dcells += '<td style="text-align:center;"><span style="' + cs + '">' + (str(n) if n > 0 else '-') + '</span></td>'
        tbody += '<tr data-pattern="' + p['pattern'] + '" data-form="' + p['form'] + '" data-total="' + str(p['total']) + '"><td><strong>' + name + '</strong></td><td style="color:#666;">' + p['form'] + '</td><td style="text-align:center;font-weight:700;">' + str(p['total']) + '</td>' + dcells + '<td><span style="font-size:11px;background:' + pbg + ';color:' + pclr + ';padding:2px 8px;border-radius:8px;font-weight:700;">' + pl + '</span></td></tr>'

    form_opts = '<option value="">All Forms</option>' + ''.join('<option value="' + f + '">' + f + '</option>' for f in sorted(forms_set))
    lbl = upload['label'] if upload else ''

    page = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Day Analysis</title><style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Inter',system-ui,-apple-system,sans-serif;background:#F3F6FB;color:#1A2638;font-size:13px;line-height:1.5;}}
.topbar{{background:#1A2638;color:white;padding:0 20px;height:52px;display:flex;align-items:center;gap:16px;border-bottom:1px solid rgba(255,255,255,0.08);}}
.topbar a{{color:rgba(255,255,255,0.6);text-decoration:none;font-size:12px;font-weight:500;}}
.topbar a:hover{{color:white;}}
.topbar h2{{font-size:14px;font-weight:700;letter-spacing:-0.2px;}}
.content{{padding:20px;max-width:1200px;margin:0 auto;}}
.card{{background:white;border-radius:12px;padding:20px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,0.06),0 1px 2px rgba(0,0,0,0.04);border:1px solid #E8EDF3;}}
.card h3{{font-size:13px;font-weight:700;margin-bottom:14px;color:#64748B;text-transform:uppercase;letter-spacing:0.5px;}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;}}
.bar-row{{display:flex;align-items:center;gap:10px;margin-bottom:10px;}}
.bar-label{{width:36px;font-size:12px;font-weight:700;}}
.bar-track{{flex:1;background:#F1F5F9;border-radius:6px;height:30px;overflow:hidden;}}
.bar-fill{{height:100%;border-radius:6px;display:flex;align-items:center;padding-left:10px;color:white;font-weight:700;font-size:12px;min-width:24px;transition:width 0.6s ease;}}
.bar-count{{width:36px;text-align:right;font-size:12px;font-weight:700;}}
.pat-card{{border-radius:10px;padding:14px;border-left:3px solid;margin-bottom:10px;}}
.pat-card strong{{font-size:13px;font-weight:700;}}
.pat-card p{{font-size:12px;color:#64748B;margin-top:5px;line-height:1.6;}}
.filters{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px;align-items:center;}}
.filters select,.filters input{{padding:7px 10px;border:1.5px solid #E2E8F0;border-radius:8px;font-size:12px;font-family:inherit;background:white;color:#1A2638;}}
.filters select:focus,.filters input:focus{{outline:none;border-color:#5B8DEF;}}
.tbl-wrap{{max-height:500px;overflow-y:auto;border-radius:8px;border:1px solid #E8EDF3;}}
table{{width:100%;border-collapse:collapse;font-size:12px;}}
th{{background:#1A2638;color:white;padding:9px 12px;text-align:left;font-size:11px;font-weight:600;letter-spacing:0.3px;position:sticky;top:0;z-index:1;}}
td{{padding:8px 12px;border-bottom:1px solid #F1F5F9;}}
tr:hover td{{background:#F8FAFC;}}
.upload-form{{background:white;border-radius:12px;padding:18px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,0.06);border:1px solid #E8EDF3;}}
.modebt{{border:2px solid #E2E8F0;background:white;border-radius:8px;padding:10px;cursor:pointer;flex:1;text-align:center;font-weight:600;color:#64748B;transition:all 0.15s;}}
.modebt.on{{border-color:#166534;background:#F0FDF4;color:#166534;}}
.file-input{{border:2px dashed #CBD5E0;border-radius:10px;padding:16px;text-align:center;background:#F8FAFC;margin:10px 0;}}
.sbtn{{background:#1A2638;color:white;border:none;border-radius:8px;padding:11px;font-size:13px;font-weight:700;cursor:pointer;width:100%;margin-top:8px;transition:background 0.15s;}}
.sbtn:hover{{background:#2D3F55;}}
.reset-btn{{padding:6px 12px;border:1.5px solid #E2E8F0;border-radius:7px;cursor:pointer;font-size:12px;background:white;color:#64748B;font-family:inherit;}}
.reset-btn:hover{{border-color:#5B8DEF;color:#5B8DEF;}}
</style></head><body>
<div class="topbar"><a href="/dashboard/{uid}" target="_top">&#8592; Back to Dashboard</a><div style="width:1px;height:20px;background:rgba(255,255,255,0.15);"></div><h2>Day of Week Analysis</h2><span style="font-size:12px;color:rgba(255,255,255,0.5);margin-left:auto;">{lbl}</span></div>
<div class="content">
<div class="upload-form"><details><summary style="cursor:pointer;font-weight:700;color:#1A5C1A;font-size:13px;">+ Upload New File (click to expand)</summary><div style="margin-top:12px;"><form method="POST" action="/dayanalysis/{uid}" enctype="multipart/form-data"><div style="display:flex;gap:8px;margin-bottom:10px;"><div class="modebt on" id="bt" onclick="sel('term')">Full Term<div style="font-size:11px;color:#888;font-weight:400;">Replaces previous</div></div><div class="modebt" id="bw" onclick="sel('week')">Weekly<div style="font-size:11px;color:#888;font-weight:400;">Adds to existing</div></div></div><input type="hidden" name="period" id="pi" value="term"><div class="file-input"><input type="file" name="file" accept=".xls,.xlsx" required></div><button type="submit" class="sbtn">Analyse</button></form></div></details></div>
<div class="g2">
<div class="card"><h3>Absences by Day of Week</h3>{bars}</div>
<div class="card"><h3>Why Students Are Absent — Pattern Analysis</h3>{insights}</div>
</div>
<div class="card">
<h3>Student Absence Patterns ({total} students)</h3>
<div class="filters">
<select id="fp" onchange="flt()"><option value="">All Patterns</option><option value="friday_always">Always misses Fri</option><option value="friday_often">Often misses Fri</option><option value="monday_often">Often misses Mon</option><option value="weekend_extended">Extended weekend</option><option value="midweek">Midweek absences</option><option value="random">Spread/Medical</option></select>
<select id="ff" onchange="flt()">{form_opts}</select>
<input type="number" id="fm" placeholder="Min absences" style="width:110px;" onchange="flt()">
<span id="fc" style="color:#888;">{total} shown</span>
<button class="reset-btn" onclick="document.getElementById('fp').value='';document.getElementById('ff').value='';document.getElementById('fm').value='';flt();">Reset filters</button>
</div>
<div class="tbl-wrap"><table><thead><tr><th>Student</th><th>Form</th><th style="text-align:center;">Total</th><th style="text-align:center;color:#A4C8E8;">Mon</th><th style="text-align:center;color:#A8DFB9;">Tue</th><th style="text-align:center;color:#CE93D8;">Wed</th><th style="text-align:center;color:#FFCC80;">Thu</th><th style="text-align:center;color:#EF9A9A;">Fri</th><th>Pattern</th></tr></thead><tbody id="tb">{tbody}</tbody></table></div>
</div>
</div>
<script>
function sel(m){{document.getElementById('pi').value=m;document.getElementById('bt').className='modebt'+(m=='term'?' on':'');document.getElementById('bw').className='modebt'+(m=='week'?' on':'');}}
function flt(){{var p=document.getElementById('fp').value,f=document.getElementById('ff').value,m=parseInt(document.getElementById('fm').value)||0,n=0;document.querySelectorAll('#tb tr').forEach(function(r){{var s=(!p||r.dataset.pattern==p)&&(!f||r.dataset.form==f)&&parseInt(r.dataset.total)>=m;r.style.display=s?'':'none';if(s)n++;}});document.getElementById('fc').textContent=n+' shown';}}
</script>
</body></html>""".format(uid=upload_id, lbl=lbl, bars=bars, insights=insights, total=len([p for p in patterns.values() if p['total']>=1]), form_opts=form_opts, tbody=tbody)
    return page


@app.route('/dayanalysis/<int:upload_id>', methods=['POST'])
@login_required
@admin_required
def dayanalysis_upload(upload_id):
    """
    Handle form-POST upload of an absentee report to the Day Analysis page.
    Parses the file, enriches students with their class from the database,
    merges with existing data in weekly mode, saves to day_analysis table,
    then redirects to the GET view to display results.
    """
    if 'file' not in request.files:
        return "<h1>No file</h1>"
    f = request.files['file']
    period = request.form.get('period', 'term')
    filename = datetime.now().strftime('%Y%m%d_%H%M%S') + '_' + secure_filename(f.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)
    print(f"Day analysis file saved: {filepath}")
    new_data = parse_absentee_file(filepath)
    if not new_data:
        return "<h1>Parse failed</h1><p><a href='/dayanalysis/{0}'>Try again</a></p>".format(upload_id)

    # Enrich with form data
    db2 = get_db()
    students_db = db2.execute("SELECT name, form FROM students WHERE upload_id=?", (upload_id,)).fetchall()
    db2.close()
    student_forms = {s['name']: s['form'] for s in students_db}
    for name in new_data:
        if name in student_forms:
            new_data[name]['form'] = student_forms[name]
        else:
            surname = name.split(',')[0].strip()
            for db_name, form in student_forms.items():
                if db_name.startswith(surname + ','):
                    new_data[name]['form'] = form
                    break

    # Weekly merge
    if period == 'week':
        db3 = get_db()
        existing_row = db3.execute("SELECT day_data FROM day_analysis WHERE upload_id=?", (upload_id,)).fetchone()
        db3.close()
        if existing_row:
            existing = json.loads(existing_row['day_data'])
            for name, nd in new_data.items():
                if name in existing:
                    for day in ['Mon','Tue','Wed','Thu','Fri']:
                        existing[name][day] = existing[name].get(day,0) + nd.get(day,0)
                    existing[name]['total'] = existing[name].get('total',0) + nd.get('total',0)
                else:
                    existing[name] = nd
            new_data = existing

    # Save
    db = get_db()
    existing = db.execute("SELECT id FROM day_analysis WHERE upload_id=?", (upload_id,)).fetchone()
    if existing:
        db.execute("UPDATE day_analysis SET day_data=? WHERE upload_id=?", (json.dumps(new_data), upload_id))
    else:
        db.execute("INSERT INTO day_analysis (upload_id, day_data) VALUES (?,?)", (upload_id, json.dumps(new_data)))
    db.commit()
    db.close()

    # Redirect back to GET to show results
    from flask import redirect
    return redirect('/dayanalysis/' + str(upload_id))


@app.route('/api/dayofweek/<int:upload_id>', methods=['GET'])
@login_required
def get_dayofweek_data(upload_id):
    """Return stored day-of-week analysis data for dashboard display"""
    db = get_db()
    row = db.execute("SELECT day_data FROM day_analysis WHERE upload_id=?", (upload_id,)).fetchone()
    db.close()
    if row:
        data = json.loads(row['day_data'])
        # Filter to own class for non-admin users
        fa = getattr(current_user, 'form_access', None)
        if fa:
            data = {name: s for name, s in data.items() if s.get('form') == fa}
        return jsonify({'data': data})
    return jsonify({'data': None})

@app.route('/guide')
@login_required
def guide():
    """Render the user guide / help page."""
    return render_template('guide.html')

@app.route('/debug')
def debug():
    """
    Development diagnostic route — shows the template path, file size, and
    first 500 characters of dashboard.html. Not protected by login; remove
    or restrict this route before exposing the app publicly.
    """
    import os
    template_path = os.path.join(app.root_path, 'templates', 'dashboard.html')
    exists = os.path.exists(template_path)
    size = os.path.getsize(template_path) if exists else 0
    with open(template_path, 'r') as f:
        first_500 = f.read(500)
    return f"<pre>Template path: {template_path}\nExists: {exists}\nSize: {size} bytes\nFirst 500 chars:\n{first_500}</pre>"

# =============================================================================
# AUTO BACKUP — runs every Monday at 7 AM while the app is running
# =============================================================================

def auto_backup():
    """Copy the live database to the backups/ folder, keep last BACKUP_KEEP copies."""
    if not os.path.exists(DB_PATH):
        return
    stamp    = datetime.now().strftime('%Y-%m-%d_%H%M')
    dest     = os.path.join(BACKUP_DIR, f'attendance_backup_{stamp}.db')
    shutil.copy2(DB_PATH, dest)
    print(f'✅ Auto-backup saved → {dest}')

    # Prune old backups — keep only the most recent BACKUP_KEEP files
    backups = sorted(glob.glob(os.path.join(BACKUP_DIR, 'attendance_backup_*.db')))
    for old in backups[:-BACKUP_KEEP]:
        os.remove(old)
        print(f'🗑  Old backup removed → {old}')

scheduler = BackgroundScheduler()
scheduler.add_job(auto_backup, 'cron', day_of_week='mon', hour=7, minute=0)
scheduler.start()


# =============================================================================
# SERVER STARTUP
# =============================================================================

if __name__ == '__main__':
    import os
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'dashboard.html')
    print("\n" + "=" * 50)
    print("  Moil Primary Attendance System")
    print("  Open your browser:  http://localhost:5000")
    print("  Share with others:  http://[YOUR-IP]:5000")
    print("=" * 50)
    print(f"  Template path: {template_path}")
    print(f"  Template exists: {os.path.exists(template_path)}")
    if os.path.exists(template_path):
        print(f"  Template size: {os.path.getsize(template_path)} bytes")
    print("=" * 50 + "\n")
    # host='0.0.0.0' makes the server accessible from other computers on the network
    # debug=False for stable production use (set True only during development)
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
